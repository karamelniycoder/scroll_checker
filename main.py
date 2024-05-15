from openpyxl.styles import Color, PatternFill, Font, Alignment, Border, Side
from openpyxl import Workbook, load_workbook
from tls_client import Session
from datetime import datetime
from loguru import logger
from requests import get
from sys import stderr
from time import sleep
import os


PROXY_TYPE = "txt" # "mobile" | "txt" | mobile - использовать мобильную проксю из `MOBILE_PROXY` | txt - использовать прокси по порядку из файла proxies.txt
MOBILE_PROXY = "http://log:pass@ip:port"
PROXY_CHANGE_LINK = "https://changeip.mobileproxy.space/?proxy_key=...&format=json"

logger.remove()
logger.add(stderr, format="<white>{time:HH:mm:ss}</white> | <level>{message}</level>")


class Excel:
    def __init__(self, total_len: int):
        if not os.path.isdir('results'): os.mkdir('results')

        workbook = Workbook()
        sheet = workbook.active
        self.file_name = f'{total_len}accs_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'

        sheet['A1'] = 'EVM Address'
        sheet['B1'] = 'Reward'

        sheet.column_dimensions['A'].width = 46
        sheet.column_dimensions['B'].width = 12

        for cell in sheet._cells:
            sheet.cell(cell[0], cell[1]).font = Font(bold=True)
            sheet.cell(cell[0], cell[1]).alignment = Alignment(horizontal='center')
            sheet.cell(cell[0], cell[1]).border = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))

        workbook.save('results/'+self.file_name)


    def edit_table(self, address: str, reward: int):
        while True:
            try:
                workbook = load_workbook('results/'+self.file_name)
                sheet = workbook.active

                valid_info = [
                    address,
                    reward
                ]
                sheet.append(valid_info)

                for row_cells in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row):
                    for cell in row_cells:
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))
                        if cell.column == 2:
                            if type(cell.value) in [float, int] and cell.value > 0: rgb_color = '32CD32'
                            else: rgb_color = 'ff0f0f'
                            cell.fill = PatternFill(patternType='solid', fgColor=Color(rgb=rgb_color))

                workbook.save('results/'+self.file_name)
                return True

            except PermissionError:
                logger.warning(f'Excel | Cant save excel file, close it!')
                sleep(3)

            except Exception as err:
                logger.critical(f'Excel | Cant save excel file: {err} | {address}')
                return False


class Proxy:
    def __init__(self):
        with open('proxies.txt') as f: self.proxies = f.read().splitlines()


    def get_proxy(self):
        if PROXY_TYPE == "txt":
            self.proxies = self.proxies[1:] + [self.proxies[0]]
            logger.debug(f'[•] Proxy | Using {self.proxies[-1]}')
            return self.proxies[-1]
        elif PROXY_TYPE == "mobile":
            self.change_proxy_ip()
            return MOBILE_PROXY


    def change_proxy_ip(self):
        if PROXY_CHANGE_LINK not in ['https://changeip.mobileproxy.space/?proxy_key=...&format=json', '']:
            while True:
                try:
                    r = get(PROXY_CHANGE_LINK)
                    if 'mobileproxy' in PROXY_CHANGE_LINK and r.json().get('status') == 'OK':
                        print('')  # empty string before next acc
                        logger.debug(f'[+] Proxy | Successfully changed ip: {r.json()["new_ip"]}')
                        return True
                    elif not 'mobileproxy' in PROXY_CHANGE_LINK and r.status_code == 200:
                        print('')  # empty string before next acc
                        logger.debug(f'[+] Proxy | Successfully changed ip: {r.text}')
                        return True
                    logger.error(f'[-] Proxy | Change IP error: {r.text} | {r.status_code}')
                    sleep(10)

                except Exception as err:
                    logger.error(f'[-] Browser | Cannot get proxy: {err}')


def get_rewards(address: str, proxy: str, excel: Excel):
    try:
        session = Session(
            client_identifier="chrome_120",
            random_tls_extension_order=True
        )
        session.headers.update({
            "Referer": "https://scroll.io/",
            "Origin": "https://scroll.io",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        })
        session.proxies.update({'http': proxy, 'https': proxy})
        r = session.get(f"https://kx58j6x5me.execute-api.us-east-1.amazonaws.com/scroll/wallet-points?walletAddress={address}")

        try: points = round(r.json()[0].get("points"), 4)
        except: raise Exception(f'Couldnt parse: {r.text}')

        if points > 0: logger.success(f'[+] Scroll | {address} | {points} marks')
        else: logger.info(f'[-] Scroll | {address} | {points} marks')

    except Exception as err:
        points = str(err)
        logger.error(f'[-] Error | {address} {err}')

    finally:
        excel.edit_table(address=address, reward=points)


if __name__ == "__main__":

    with open("addresses.txt") as f: addresses = f.read().splitlines()

    proxy_manager = Proxy()

    input('\n\t> Start')

    excel = Excel(len(addresses))

    for address in addresses:
        get_rewards(address=address, proxy=proxy_manager.get_proxy(), excel=excel)

    input(f'\nResults saved in {excel.file_name}\n\n\t> Exit')
